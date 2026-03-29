"""
Plasma concentration data cleaning script.
Source: data/plasma_recordings.xlsx
Paper: Timmermann et al., 2019 — Neural correlates of the DMT experience

Produces tidy long-format CSVs with one row per blood draw:
    subject | condition | dose_mg | time_point | time_min | plasma_conc | is_imputed

Usage:
    python scripts/clean_plasma_csv.py
"""

import re
from pathlib import Path
import pandas as pd
import openpyxl


ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "data" / "plasma_recordings.xlsx"

# Excel row/col indices (1-based, as openpyxl uses)
CONC_ROWS = range(5, 18)   # rows 5–17: concentration data
TIME_ROWS = range(26, 39)  # rows 26–38: time point data

DMT_COLS = {"dose": 1, "subject": 2, "t0": 3, "t1": 4, "t2": 5, "t3": 6, "t4": 7}   # A–G
PLACEBO_COLS = {"subject": 10, "t0": 11, "t1": 12, "t2": 13, "t3": 14, "t4": 15}  # J–O


# ---------------------------------------------------------------------------
# Step 1 — dual load
# ---------------------------------------------------------------------------

wb_vals = openpyxl.load_workbook(XLSX_PATH, data_only=True)
wb_formulas = openpyxl.load_workbook(XLSX_PATH, data_only=False)
ws_vals = wb_vals["Sheet1"]
ws_fmls = wb_formulas["Sheet1"]

# Collect positions of formula cells (row, col) in the data region
formula_cells: set[tuple[int, int]] = set()
for row in ws_fmls.iter_rows(min_row=5, max_row=17, min_col=1, max_col=15):
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            formula_cells.add((cell.row, cell.column))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get(ws, row, col):
    """Return cell value (None if empty)."""
    return ws.cell(row=row, column=col).value


def is_formula(row, col):
    return (row, col) in formula_cells


def parse_dose(label: str) -> int:
    """'7 mg' → 7"""
    m = re.search(r"\d+", str(label))
    if not m:
        raise ValueError(f"Cannot parse dose from: {label!r}")
    return int(m.group())


# ---------------------------------------------------------------------------
# Step 2+3 — parse concentrations and times into dicts keyed by subject
# ---------------------------------------------------------------------------

def parse_block(ws, conc_rows, time_rows, subject_col, dose_col, conc_t_cols, time_t_cols):
    """
    Returns a list of dicts:
        {subject, dose_label, c0..c4, t0..t4, imputed_tp: set of time_point indices}
    dose_col=None for placebo.
    """
    conc_t_cols = list(conc_t_cols)   # [col_t0, col_t1, ..., col_t4]
    time_t_cols = list(time_t_cols)

    # --- concentrations ---
    records = []
    last_dose = None
    for row in conc_rows:
        subj = get(ws, row, subject_col)
        if subj is None:
            continue

        if dose_col is not None:
            raw_dose = get(ws, row, dose_col)
            if raw_dose is not None:
                last_dose = raw_dose
            dose_label = last_dose
        else:
            dose_label = None

        imputed = set()
        concs = []
        for tp, col in enumerate(conc_t_cols):
            concs.append(get(ws, row, col))
            if is_formula(row, col):
                imputed.add(tp)

        records.append({
            "subject": str(subj),
            "dose_label": dose_label,
            "concs": concs,
            "imputed_tps": imputed,
        })

    # --- times ---
    times_by_subject: dict[str, list] = {}
    for row in time_rows:
        subj = get(ws, row, subject_col)
        if subj is None:
            continue
        times_by_subject[str(subj)] = [get(ws, row, col) for col in time_t_cols]

    # --- merge ---
    rows = []
    for rec in records:
        subj = rec["subject"]
        times = times_by_subject.get(subj)
        if times is None:
            raise KeyError(f"No time data found for subject {subj!r}")
        rec["times"] = times
        rows.append(rec)

    return rows


dmt_records = parse_block(
    ws_vals,
    conc_rows=CONC_ROWS,
    time_rows=TIME_ROWS,
    subject_col=DMT_COLS["subject"],
    dose_col=DMT_COLS["dose"],
    conc_t_cols=[DMT_COLS[f"t{i}"] for i in range(5)],
    time_t_cols=[DMT_COLS[f"t{i}"] for i in range(5)],
)

placebo_records = parse_block(
    ws_vals,
    conc_rows=CONC_ROWS,
    time_rows=TIME_ROWS,
    subject_col=PLACEBO_COLS["subject"],
    dose_col=None,
    conc_t_cols=[PLACEBO_COLS[f"t{i}"] for i in range(5)],
    time_t_cols=[PLACEBO_COLS[f"t{i}"] for i in range(5)],
)

# ---------------------------------------------------------------------------
# Step 4+5 — melt each condition to long format
# ---------------------------------------------------------------------------


def melt_records(records, condition: str) -> pd.DataFrame:
    rows = []
    for rec in records:
        subj = rec["subject"]
        dose_label = rec["dose_label"]
        dose_mg = parse_dose(dose_label) if dose_label is not None else pd.NA

        for tp in range(5):
            rows.append({
                "subject": subj,
                "condition": condition,
                "dose_mg": dose_mg,
                "time_point": tp,
                "time_min": float(rec["times"][tp]),
                "plasma_conc": float(rec["concs"][tp]),
                "is_imputed": tp in rec["imputed_tps"],
            })

    return pd.DataFrame(rows)


dmt_long = melt_records(dmt_records, condition="dmt")
placebo_long = melt_records(placebo_records, condition="placebo")


# ---------------------------------------------------------------------------
# Step 6 — dose_mg type + round imputed floats
# ---------------------------------------------------------------------------

# Cast dose_mg to nullable Int64
dmt_long["dose_mg"] = dmt_long["dose_mg"].astype("Int64")
placebo_long["dose_mg"] = placebo_long["dose_mg"].astype("Int64")  # all NA


# ---------------------------------------------------------------------------
# Step 7 — concatenate & validate
# ---------------------------------------------------------------------------

df = pd.concat([dmt_long, placebo_long], ignore_index=True)

# Round imputed floating-point noise
df.loc[df["is_imputed"], "plasma_conc"] = df.loc[df["is_imputed"], "plasma_conc"].round(4)
df.loc[df["is_imputed"], "time_min"] = df.loc[df["is_imputed"], "time_min"].round(4)

# Assertions
assert len(df) == 130, f"Expected 130 rows, got {len(df)}"
assert df["plasma_conc"].isna().sum() == 0, "NaN in plasma_conc"
assert df["time_min"].isna().sum() == 0, "NaN in time_min"
assert (df["time_min"] >= 0).all(), "Negative time_min found"
assert set(df["subject"].unique()) == {f"S{i:02d}" for i in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]}
assert set(df["condition"].unique()) == {"dmt", "placebo"}

print(f"Rows: {len(df)}")
print(f"Imputed: {df['is_imputed'].sum()}")
print(f"Subjects: {sorted(df['subject'].unique())}")
print(df.head(10).to_string(index=False))


# ---------------------------------------------------------------------------
# Step 8 — export
# ---------------------------------------------------------------------------

df.to_csv(ROOT / "data" / "plasma_clean.csv", index=False)
df.to_parquet(ROOT / "data" / "plasma_clean.parquet", index=False)

print("\nSaved: data/plasma_clean.csv, data/plasma_clean.parquet")
